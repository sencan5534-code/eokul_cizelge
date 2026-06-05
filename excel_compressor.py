from openpyxl import load_workbook
from openpyxl.styles import Border
from copy import copy



TEMPLATE_SHEET = "Template"
START_ROW_TEMPLATE = 3

sn_col = 1
okul_no_col = 2
ad_soyad_col = 3

criteria_cols = {
    "Derse gerektiği kıyafetle katılır": 4,
    "Sınıf içi etkinliklere katılır": 5,
    "Arkadaşlarıyla işbirliği içinde çalışır": 6,
    "Derse karşı ilgilidir": 7,
    "Öğretmenine ve arkadaşlarına  saygılıdır": 8
}

total_col = 9
CRITERIA = list(criteria_cols.keys())

PERF_COLUMNS = {
    "1.Perf": "I",
    "2.Perf": "J"
}


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def first_not_empty(*values):
    for value in values:
        if value is not None and str(value).strip() != "":
            return value
    return None


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def copy_row_from_template(template_ws, target_ws, source_row, target_row):
    for col in range(1, template_ws.max_column + 1):
        src = template_ws.cell(source_row, col)
        dst = target_ws.cell(target_row, col)

        dst.value = src.value

        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)

    target_ws.row_dimensions[target_row].height = (
        template_ws.row_dimensions[source_row].height
    )


def calculate_scores(perf_value, status_value=None):
    status = normalize_text(status_value)

    if status == "g":
        return [0, 0, 0, 0, 0], 0

    if perf_value is None or str(perf_value).strip() == "":
        return [0, 0, 0, 0, 0], 0

    try:
        total = int(float(str(perf_value).replace(",", ".")))
    except ValueError:
        return [0, 0, 0, 0, 0], 0

    if total <= 0:
        return [0, 0, 0, 0, 0], 0

    total = max(0, min(total, 100))

    if total >= 100:
        scores = [20, 20, 20, 20, 20]
    elif 80 <= total <= 99:
        scores = [total - 80, 20, 20, 20, 20]
    elif 60 <= total <= 79:
        scores = [0, total - 60, 20, 20, 20]
    elif 40 <= total <= 59:
        scores = [0, 0, total - 40, 20, 20]
    elif 20 <= total <= 39:
        scores = [0, 0, 0, total - 20, 20]
    else:
        scores = [0, 0, 0, 0, total]

    return scores, total


def get_merged_cell_value(ws, cell_ref):
    cell = ws[cell_ref]

    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            ).value

    return None


def find_template_footer_rows(template_ws):
    last_number_row = None

    for row in range(START_ROW_TEMPLATE, template_ws.max_row + 1):
        value = template_ws.cell(row, 1).value

        if value is None:
            continue

        try:
            int(value)
            last_number_row = row
        except:
            pass

    if last_number_row is None:
        raise Exception("Template sheetinde A kolonunda sıra numarası bulunamadı.")

    return [last_number_row + 1, last_number_row + 2]


def create_perf_sheet(
    wb,
    template_ws,
    ws_source,
    source_sheet_name,
    perf_name,
    perf_col
):
    new_sheet_name = f"{source_sheet_name} {perf_name} list"

    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    ws_template = wb.copy_worksheet(template_ws)
    ws_template.title = new_sheet_name

    footer_rows = find_template_footer_rows(template_ws)

    for row in range(START_ROW_TEMPLATE, ws_template.max_row + 1):
        for col in range(1, ws_template.max_column + 1):
            ws_template.cell(row, col).value = None
            ws_template.cell(row, col).border = Border()

    target_row = START_ROW_TEMPLATE
    sira_no = 1

    for row in range(2, ws_source.max_row + 1, 2):
        row1 = row
        row2 = row + 1

        okul_no = first_not_empty(
            get_merged_cell_value(ws_source, f"A{row1}"),
            get_merged_cell_value(ws_source, f"A{row2}")
        )

        ad_soyad = first_not_empty(
            get_merged_cell_value(ws_source, f"B{row1}"),
            get_merged_cell_value(ws_source, f"B{row2}")
        )

        status = first_not_empty(
            get_merged_cell_value(ws_source, f"G{row1}"),
            get_merged_cell_value(ws_source, f"G{row2}")
        )

        perf = first_not_empty(
            get_merged_cell_value(ws_source, f"{perf_col}{row1}"),
            get_merged_cell_value(ws_source, f"{perf_col}{row2}")
        )

        if okul_no is None and ad_soyad is None:
            continue

        copy_row_style(ws_template, START_ROW_TEMPLATE, target_row)

        scores, total = calculate_scores(perf, status)

        ws_template.cell(target_row, sn_col).value = sira_no
        ws_template.cell(target_row, okul_no_col).value = okul_no
        ws_template.cell(target_row, ad_soyad_col).value = ad_soyad

        for criterion_name, score in zip(CRITERIA, scores):
            ws_template.cell(
                target_row,
                criteria_cols[criterion_name]
            ).value = score

        ws_template.cell(target_row, total_col).value = total

        target_row += 1
        sira_no += 1

    # Öğrenci tablosu tamamen bittikten sonra border uygula
    template_border_row = START_ROW_TEMPLATE

    for row in range(START_ROW_TEMPLATE, target_row):
        for col in range(1, total_col + 1):
            ws_template.cell(row, col).border = copy(
                template_ws.cell(template_border_row, col).border
            )

    # Template footer 2 satır olarak eklenecek.
    # Footer'ın ilk satırında D kolonuna yeni sheet adı bold yazılacak.
    for index, footer_source_row in enumerate(footer_rows):
        copy_row_from_template(
            template_ws,
            ws_template,
            footer_source_row,
            target_row
        )

        if index == 0:
            ws_template.cell(target_row, 4).value = new_sheet_name
            ws_template.cell(target_row, 4).font = copy(
                template_ws.cell(footer_source_row, 4).font
            )
            current_font = copy(ws_template.cell(target_row, 4).font)
            current_font.bold = True
            ws_template.cell(target_row, 4).font = current_font

        target_row += 1

    # Footer sonrasındaki satırları temizle
    empty_border = Border()

    for row in range(target_row, ws_template.max_row + 1):
        for col in range(1, ws_template.max_column + 1):
            ws_template.cell(row, col).value = None
            ws_template.cell(row, col).border = empty_border

    # Ad Soyad kolon genişliği otomatik ayarla
    max_length = 0

    for row in range(START_ROW_TEMPLATE, target_row + 1):
        cell_value = ws_template.cell(row, 3).value

        if cell_value:
            max_length = max(max_length, len(str(cell_value)))

    ws_template.column_dimensions["C"].width = max_length + 5


def process_excel(input_file, output_file):
    wb = load_workbook(input_file)
    template_ws = wb[TEMPLATE_SHEET]

    source_sheets = [
        s for s in wb.sheetnames
        if s != TEMPLATE_SHEET
        and "1.Perf list" not in s
        and "2.Perf list" not in s
        and not s.endswith(" list")
    ]

    for source_sheet_name in source_sheets:
        ws_source = wb[source_sheet_name]

        for perf_name, perf_col in PERF_COLUMNS.items():
            create_perf_sheet(
                wb,
                template_ws,
                ws_source,
                source_sheet_name,
                perf_name,
                perf_col
            )

    for sheet_name in wb.sheetnames[:]:
        if not (
            "1.Perf list" in sheet_name
            or "2.Perf list" in sheet_name
        ):
            del wb[sheet_name]

    wb.save(output_file)

    return output_file
